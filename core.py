"""core.py: Core functions of Web Scraper used in passExcel, GUI and CLI versions of Test_Carte."""

__author__ = "Arnaud Petit"

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import os
import time
from openpyxl import load_workbook, Workbook  #excel tools
from datetime import date

##ini
from backports import configparser

config = configparser.ConfigParser()

delay = 5  #timeout in seconds


def creation_ini():
    f = open("config.ini", "x")
    myFile = open('config.ini', 'w')
    myFile.write(
        '[outputConfig]\noutputRepertory =\ninputRepertory=\n[prodSercelLogin]\nid =\npassword='
    )
    myFile.close()


def repertory():
    config = configparser.ConfigParser()
    if not (os.path.exists("config.ini")):
        creation_ini()

    config.read("config.ini")
    destinationRepertory = config["outputConfig"].get("outputRepertory")
    incomingRepertory = config["outputConfig"].get("inputRepertory")
    ##output
    if destinationRepertory:
        outputRepertory = destinationRepertory
        if outputRepertory[-1] != "\\":
            outputRepertory += "\\"
    else:  #output : Desktop/Out
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']),
                               'Desktop')  #desktop path
        outputRepertory = desktop + '\\Out\\'  #Out Folder
    ##input
    if incomingRepertory:
        filesRepertory = incomingRepertory
        if filesRepertory[-1] != "\\":
            filesRepertory += "\\"
    else:  #output : Desktop/Out
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']),
                               'Desktop')  #desktop path
        filesRepertory = desktop + '\\In\\'  #Out Folder

    if not (os.path.exists(outputRepertory)):
        os.mkdir(outputRepertory[:-1])

    if not (os.path.exists(filesRepertory)):
        os.mkdir(filesRepertory[:-1])

    return filesRepertory, outputRepertory


def initialisation():
    #caché

    global driver
    global config
    if not (os.path.exists("config.ini")):
        creation_ini()

    config.read("config.ini")
    login = str(config["prodSercelLogin"].get("id"))
    password = str(config["prodSercelLogin"].get("password"))
    if not (login and password):  #arthur's credentials as fallback
        login = "GITHUB_DELETED"
        password = "GITHUB_DELETED"

    #prevent for error message on start (Sercel has an exotic safety rule on navigators)
    args = [
        "hide_console",
    ]
    chromeOptions = webdriver.ChromeOptions()
    chromeOptions.add_experimental_option('useAutomationExtension', False)

    driver = webdriver.Chrome(
        options=chromeOptions,
        desired_capabilities=chromeOptions.to_capabilities(),
        service_args=args)
    #hide the working window by sending it very far
    driver.set_window_position(-10000, 0)

    #login as Arthur
    driver.get("http://prod.sercel.fr/tonicweb/index.php")
    driver.find_element_by_id("UserRememberMe").click()
    driver.find_element_by_id("UserUsername").send_keys(login)
    driver.find_element_by_id("UserPassword").send_keys(password +
                                                        "\n")  #\n to valid

    #check if we are logged in correctly
    try:
        if (driver.find_element_by_class_name("message").get_attribute(
                "innerText") == "Nom d'utilisateur ou mot de passe incorrect."):

            return "loginFail"

    except:
        pass


def removeZero(string):

    i = 0
    while string[i] == "0":
        i += 1
    return string[i:]


def test(rawInput):  #return bool (pass/fail) or str (not enough noise test)

    def countOccurence(string, bigString):
        length = len(string)
        counter = 0
        for i in range(len(bigString) - length + 1):
            if bigString[i:i + length] == string:
                counter += 1
        return counter

    urlIntranet = "http://prod.sercel.fr/tonicweb/index.php/uuts/view/uut_num:" + removeZero(
        rawInput)
    driver.get(urlIntranet)
    try:
        WebDriverWait(driver, delay).until(
            EC.presence_of_element_located((By.ID, 'line5')))

        classFirstRow = driver.find_element_by_css_selector(
            "tr:nth-of-type(2)").get_attribute("class")

        if not ("Noise_Pre-Assembly" in driver.find_element_by_css_selector(
                "tr:nth-of-type(2) td:nth-of-type(1) a").get_attribute(
                    'innerHTML')):  #last test is not noise
            return "unexpected"

        if (classFirstRow == 'fail' or classFirstRow == 'alt-fail'
           ):  #last test : supposed to be a noise one -> is fail
            listeTest = ""
            continuer = True
            i = 2  #start with second row
            while continuer:
                try:
                    listeTest += driver.find_element_by_css_selector(
                        "tr:nth-of-type({}) td:nth-of-type(1) a".format(
                            i)).get_attribute('innerHTML')
                except:
                    continuer = False
                i += 1

            #listeTest contains all the test not separated : optimal for search

            if countOccurence("Noise_Pre-Assembly",
                              listeTest) < 6:  #not enough test
                return ("lackOfTest")

            #fail case, let's get the first error occured on last test
            driver.find_element_by_css_selector(
                "tr:nth-of-type(2) td:nth-of-type(12) a").click()
            #WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.ID, 'gridbox')))
            WebDriverWait(
                driver, delay).until(lambda d: d.find_element_by_tag_name("td"))
            time.sleep(
                0.1
            )  #needed to predict latence after first table data (td) appears
            chaineErreurs = ""
            continuer = True
            i = 1
            while continuer:
                try:
                    chaineErreurs += "|" + driver.find_element_by_xpath(
                        "(//*[contains(@style, 'background-color: rgb(255, 128, 128);')]/..)[{}]"
                        .format(i)).find_element_by_css_selector(
                            "td:nth-of-type(5) a").get_attribute(
                                "innerHTML").replace("&amp;", "&")
                except:
                    continuer = False
                i += 1

            return "fail " + chaineErreurs[1:]

        elif (classFirstRow == 'abort' or classFirstRow == 'alt-abort'):  #abort
            return ("abort")

        return "pass"  #eveything is fine

    except:  #error on webdriverWait : nothing found on page -> incorrect serial number
        return "fail"


def done():
    global driver
    driver.close()


def enregistrer(serial, res):
    inputDirectory, outputRepertory = repertory()

    dailySheetName = date.today().strftime("%d_%m_%Y") + "_Scan.xlsx"

    #excel save feature
    if not (os.path.exists(outputRepertory +
                           dailySheetName)):  #touch daily file if not exists
        wbDouche = Workbook()
        sheetRanges = wbDouche.active
        indice = 2
        sheetRanges['A1'].value = "Serial Number"
        sheetRanges['B1'].value = "TonicWeb"
        sheetRanges['C1'].value = "Cause of Failure"
    else:
        wbDouche = load_workbook(filename=outputRepertory + dailySheetName)
        sheetRanges = wbDouche.active
        indice = 1
        while sheetRanges['A' + str(indice)].value:  #count rows
            indice += 1

    sheetRanges['A' + str(indice)].value = serial

    if res[:4] != "fail":
        sheetRanges['B' + str(indice)].value = res

    else:
        sheetRanges['B' + str(indice)].value = res[:4]
        sheetRanges['C' + str(indice)].value = res[5:]

    wbDouche.save(outputRepertory + dailySheetName)
